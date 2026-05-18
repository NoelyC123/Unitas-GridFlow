"""Excel and shared survey export logic."""

from __future__ import annotations

import csv
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

from gridflow.conflict_detector import ConflictDetector
from gridflow.evidence_combiner import EvidenceCombiner, combine_pole_evidence, link_pole
from gridflow.photos import load_pole_photos
from gridflow.readiness import ReadinessAssessor

POLES_COLUMNS = [
    "pole_folder",
    "support_no",
    "pole_type",
    "linking_confidence",
    "linking_method",
    "enwl_asset_type",
    "enwl_voltage",
    "enwl_circuit_name",
    "conflict_count",
    "critical_conflicts",
    "photo_count",
    "photo_types_present",
    "readiness_status",
    "design_ready",
    "readiness_confidence",
    "readiness_blockers",
    "readiness_warnings",
    "notes_excerpt",
]


@dataclass
class ExportBundle:
    survey_name: str
    export_timestamp: str
    trace_source_file: str
    pole_rows: list[dict[str, Any]]
    conflict_rows: list[dict[str, Any]]
    photo_rows: list[dict[str, Any]]
    total_photos: int

    @property
    def total_poles(self) -> int:
        return len(self.pole_rows)

    @property
    def ready_count(self) -> int:
        return sum(1 for row in self.pole_rows if row["readiness_status"] == "ready")

    @property
    def review_required_count(self) -> int:
        return sum(1 for row in self.pole_rows if row["readiness_status"] == "review_required")

    @property
    def not_ready_count(self) -> int:
        return sum(1 for row in self.pole_rows if row["readiness_status"] == "not_ready")

    @property
    def insufficient_evidence_count(self) -> int:
        return sum(
            1 for row in self.pole_rows if row["readiness_status"] == "insufficient_evidence"
        )

    @property
    def total_conflicts(self) -> int:
        return len(self.conflict_rows)


class SurveyExportCollector:
    """Collect flat export records from survey evidence and optional trace data."""

    def __init__(self) -> None:
        self._combiner = EvidenceCombiner()
        self._conflict_detector = ConflictDetector()
        self._readiness_assessor = ReadinessAssessor()

    def collect(self, survey_root: str | Path, trace_path: str | Path) -> ExportBundle:
        survey_root = Path(survey_root)
        poles_root = survey_root / "enwl_enrichment_clean"
        if not poles_root.exists():
            raise FileNotFoundError(f"Survey evidence folder not found: {poles_root}")

        pole_dirs = sorted(
            [path for path in poles_root.iterdir() if path.is_dir() and "_SUPPORT_" in path.name],
            key=lambda path: _pole_sort_key(path.name),
        )
        if not pole_dirs:
            raise ValueError(f"Survey contains no pole folders: {poles_root}")

        trace_path = Path(trace_path)
        trace_exists = trace_path.exists()

        pole_rows: list[dict[str, Any]] = []
        conflict_rows: list[dict[str, Any]] = []
        photo_rows: list[dict[str, Any]] = []
        total_photos = 0

        for pole_dir in pole_dirs:
            photo_set = load_pole_photos(pole_dir)
            notes_path = pole_dir / "notes" / "pole_notes.md"
            notes_excerpt = _notes_excerpt(notes_path)
            parsed_notes = (
                self._combiner.parse_pole_notes(notes_path)
                if notes_path.exists()
                else self._empty_notes(notes_path)
            )
            total_photos += photo_set.photo_count

            if trace_exists:
                combined = combine_pole_evidence(survey_root, pole_dir.name, trace_path)
                linking = link_pole(survey_root, pole_dir.name, trace_path)
                conflicts = self._conflict_detector.detect_pole(
                    survey_root, pole_dir.name, trace_path
                )
                readiness = self._readiness_assessor.assess_from_records(
                    combined, linking, conflicts
                )
            else:
                combined = self._fallback_combined(pole_dir, parsed_notes, photo_set, notes_path)
                linking = _fallback_linking(pole_dir.name, parsed_notes)
                conflicts = []
                readiness = self._readiness_assessor.assess_from_records(
                    combined, linking, conflicts
                )

            pole_row = self._pole_row_from_records(
                pole_dir.name,
                combined,
                linking,
                readiness,
                photo_set,
                notes_excerpt,
            )
            pole_rows.append(pole_row)

            counts = photo_set.count_by_type()
            photo_rows.append(
                {
                    "pole_folder": pole_dir.name,
                    "support_no": combined.get("support_no") or parsed_notes.support_no or "",
                    "total_photos": photo_set.photo_count,
                    "full_pole": counts["full_pole"],
                    "pole_top": counts["pole_top"],
                    "pole_base": counts["pole_base"],
                    "equipment": counts["equipment"],
                    "span": counts["span"],
                    "context": counts["context"],
                    "unknown": counts["unknown"],
                }
            )

            for conflict in conflicts:
                conflict_rows.append(
                    {
                        "pole_folder": pole_dir.name,
                        "support_no": combined.get("support_no") or parsed_notes.support_no or "",
                        "conflict_type": conflict.conflict_type,
                        "severity": conflict.severity,
                        "field_value": conflict.field_value or "",
                        "enwl_value": conflict.enwl_value or "",
                        "description": conflict.description,
                    }
                )

        return ExportBundle(
            survey_name=survey_root.name,
            export_timestamp=datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            trace_source_file=str(trace_path),
            pole_rows=pole_rows,
            conflict_rows=conflict_rows,
            photo_rows=photo_rows,
            total_photos=total_photos,
        )

    def _fallback_combined(
        self,
        pole_dir: Path,
        parsed_notes: Any,
        photo_set: Any,
        notes_path: Path,
    ) -> dict[str, Any]:
        return {
            "pole_id": pole_dir.name,
            "support_no": parsed_notes.support_no,
            "pole_fid": parsed_notes.pole_fid,
            "spn": parsed_notes.spn,
            "pole_type": parsed_notes.pole_type,
            "pole_class": parsed_notes.pole_class,
            "support_diameter": parsed_notes.support_diameter,
            "coordinates": parsed_notes.coordinates,
            "photo_count": photo_set.photo_count,
            "photo_types_present": photo_set.photo_types_present,
            "photos_available": photo_set.photo_count > 0,
            "direct_equipment_records": [],
            "route_conductor_evidence": [],
            "nearby_context": [],
            "evidence_quality_summary": {
                "pole_identity": "HIGH"
                if parsed_notes.support_no and parsed_notes.pole_fid and parsed_notes.spn
                else "LOW",
                "direct_equipment": "NONE",
                "route_conductor": "NONE",
                "nearby_context": "NONE",
                "direct_equipment_count": 0,
                "route_conductor_count": 0,
                "nearby_context_count": 0,
            },
            "design_readiness_caution": (
                "Trace data unavailable during export; readiness and conflict fields are conservative defaults."
            ),
            "contributing_files": {
                "survey_root": str(pole_dir.parents[1]),
                "pole_folder": str(pole_dir),
                "field_photos": str(pole_dir / "field_photos"),
                "pole_notes": str(notes_path),
                "trace_geojson": "",
            },
            "uncertainties": list(parsed_notes.uncertainties),
        }

    def _empty_notes(self, notes_path: Path) -> Any:
        return (
            self._combiner.parse_pole_notes(notes_path)
            if notes_path.exists()
            else type(
                "EmptyNotes",
                (),
                {
                    "support_no": None,
                    "pole_fid": None,
                    "spn": None,
                    "pole_type": None,
                    "pole_class": None,
                    "support_diameter": None,
                    "coordinates": None,
                    "uncertainties": [],
                },
            )()
        )

    def _pole_row_from_records(
        self,
        pole_folder: str,
        combined: dict[str, Any],
        linking: Any,
        readiness: Any,
        photo_set: Any,
        notes_excerpt: str,
    ) -> dict[str, Any]:
        enwl_voltage = (
            _first_nonempty(
                *[record.get("voltage") for record in combined.get("direct_equipment_records", [])],
                *[record.get("voltage") for record in combined.get("route_conductor_evidence", [])],
                *[record.get("voltage") for record in combined.get("nearby_context", [])],
            )
            or ""
        )
        enwl_circuit_name = (
            _first_nonempty(
                *[record.get("text_map") for record in combined.get("route_conductor_evidence", [])]
            )
            or ""
        )
        return {
            "pole_folder": pole_folder,
            "support_no": combined.get("support_no") or "",
            "pole_type": combined.get("pole_type") or "",
            "linking_confidence": getattr(linking, "confidence", "NONE"),
            "linking_method": getattr(linking, "linking_method", "none"),
            "enwl_asset_type": combined.get("pole_type") or "",
            "enwl_voltage": enwl_voltage,
            "enwl_circuit_name": enwl_circuit_name,
            "conflict_count": readiness.conflict_count,
            "critical_conflicts": readiness.critical_conflicts,
            "photo_count": photo_set.photo_count,
            "photo_types_present": ", ".join(photo_set.photo_types_present),
            "readiness_status": readiness.readiness_status,
            "design_ready": readiness.design_ready,
            "readiness_confidence": readiness.readiness_confidence,
            "readiness_blockers": "; ".join(readiness.readiness_blockers),
            "readiness_warnings": "; ".join(readiness.readiness_warnings),
            "notes_excerpt": notes_excerpt,
        }


class SurveyExcelExporter:
    """Export survey evidence to a formatted Excel workbook."""

    def __init__(self) -> None:
        self._collector = SurveyExportCollector()

    def export(
        self, survey_root: str | Path, trace_path: str | Path, output_path: str | Path
    ) -> ExportBundle:
        from openpyxl import Workbook

        bundle = self._collector.collect(survey_root, trace_path)

        workbook = Workbook()
        summary_sheet = workbook.active
        summary_sheet.title = "Summary"
        poles_sheet = workbook.create_sheet("Poles")
        conflicts_sheet = workbook.create_sheet("Conflicts")
        photos_sheet = workbook.create_sheet("Photos")

        self._write_summary(summary_sheet, bundle)
        self._write_poles(poles_sheet, bundle.pole_rows)
        self._write_conflicts(conflicts_sheet, bundle.conflict_rows)
        self._write_photos(photos_sheet, bundle.photo_rows)

        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(output_path)
        return bundle

    def _write_summary(self, sheet, bundle: ExportBundle) -> None:
        rows = [
            ("Survey name", bundle.survey_name),
            ("Export date/time", bundle.export_timestamp),
            ("Total poles", bundle.total_poles),
            ("Poles ready (design_ready=true)", bundle.ready_count),
            ("Poles review_required", bundle.review_required_count),
            ("Poles not_ready", bundle.not_ready_count),
            ("Poles insufficient_evidence", bundle.insufficient_evidence_count),
            ("Total photos detected", bundle.total_photos),
            ("Total conflicts detected", bundle.total_conflicts),
            ("Trace source file", bundle.trace_source_file),
        ]
        sheet.append(["Metric", "Value"])
        for metric, value in rows:
            sheet.append([metric, value])
        self._style_header(sheet)
        self._freeze(sheet)
        self._autowidth(sheet)

    def _write_poles(self, sheet, rows: list[dict[str, Any]]) -> None:
        sheet.append(POLES_COLUMNS)
        for index, row in enumerate(rows, start=2):
            sheet.append([self._excel_value(row.get(column)) for column in POLES_COLUMNS])
            fill = _status_fill(row["readiness_status"], index)
            for cell in sheet[index]:
                cell.fill = fill
        self._style_header(sheet)
        self._freeze(sheet)
        self._autowidth(sheet)

    def _write_conflicts(self, sheet, rows: list[dict[str, Any]]) -> None:
        columns = [
            "pole_folder",
            "support_no",
            "conflict_type",
            "severity",
            "field_value",
            "enwl_value",
            "description",
        ]
        sheet.append(columns)
        for row in rows:
            sheet.append([self._excel_value(row.get(column)) for column in columns])
        self._style_header(sheet)
        self._freeze(sheet)
        self._autowidth(sheet)

    def _write_photos(self, sheet, rows: list[dict[str, Any]]) -> None:
        columns = [
            "pole_folder",
            "support_no",
            "total_photos",
            "full_pole",
            "pole_top",
            "pole_base",
            "equipment",
            "span",
            "context",
            "unknown",
        ]
        sheet.append(columns)
        for row in rows:
            sheet.append([self._excel_value(row.get(column)) for column in columns])
        self._style_header(sheet)
        self._freeze(sheet)
        self._autowidth(sheet)

    def _style_header(self, sheet) -> None:
        from openpyxl.styles import Font

        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.fill = _pattern_fill("D9EAF7")

    def _freeze(self, sheet) -> None:
        sheet.freeze_panes = "A2"

    def _autowidth(self, sheet) -> None:
        for column_cells in sheet.columns:
            length = max(len(str(cell.value or "")) for cell in column_cells)
            sheet.column_dimensions[column_cells[0].column_letter].width = min(
                max(length + 2, 12), 50
            )

    def _excel_value(self, value: Any) -> Any:
        if isinstance(value, bool):
            return "TRUE" if value else "FALSE"
        return value


class SurveyCSVExporter:
    """Export survey evidence to a flat CSV."""

    def __init__(self) -> None:
        self._collector = SurveyExportCollector()

    def export(
        self, survey_root: str | Path, trace_path: str | Path, output_path: str | Path
    ) -> ExportBundle:
        bundle = self._collector.collect(survey_root, trace_path)
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        with output_path.open("w", encoding="utf-8", newline="") as handle:
            writer = csv.DictWriter(handle, fieldnames=POLES_COLUMNS)
            writer.writeheader()
            for row in bundle.pole_rows:
                writer.writerow({key: self._csv_value(row.get(key)) for key in POLES_COLUMNS})
        return bundle

    def _csv_value(self, value: Any) -> Any:
        if isinstance(value, bool):
            return "TRUE" if value else "FALSE"
        return value


def _notes_excerpt(notes_path: Path) -> str:
    if not notes_path.exists():
        return ""
    text = " ".join(notes_path.read_text(encoding="utf-8").split())
    return text[:200]


def _first_nonempty(*values: Any) -> Any:
    for value in values:
        if value not in (None, ""):
            return value
    return None


def _pole_sort_key(pole_id: str) -> tuple[int, str]:
    prefix = pole_id.split("_", 1)[0]
    return (int(prefix) if prefix.isdigit() else 999, pole_id)


def _fallback_linking(pole_id: str, parsed_notes: Any):
    return type(
        "FallbackLinking",
        (),
        {
            "pole_id": pole_id,
            "support_no": parsed_notes.support_no,
            "pole_fid": parsed_notes.pole_fid,
            "linking_method": "none",
            "confidence": "NONE",
            "matched_enwl_fid": parsed_notes.pole_fid,
            "evidence_source": "pole_notes_only",
            "manual_confirmation_required": True,
            "distance_m": None,
            "notes": "Trace path missing during export.",
            "matched_methods": [],
            "direct_equipment_fids": [],
            "matched_support_no": None,
            "matched_spn": None,
        },
    )()


def _status_fill(status: str, row_number: int):

    odd = row_number % 2 == 1
    if status == "ready":
        return _pattern_fill("DFF2E1" if odd else "D4EBD6")
    if status == "review_required":
        return _pattern_fill("FFF4D6" if odd else "FCE9BF")
    if status == "not_ready":
        return _pattern_fill("FADDDD" if odd else "F3CFCF")
    if status == "insufficient_evidence":
        return _pattern_fill("E7E7E7" if odd else "DADADA")
    return _pattern_fill("F8FBFE" if odd else "EEF5FB")


def _pattern_fill(color: str):
    from openpyxl.styles import PatternFill

    return PatternFill("solid", fgColor=color)
