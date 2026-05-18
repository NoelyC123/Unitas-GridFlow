from __future__ import annotations

import csv
import json
import subprocess
from pathlib import Path

from gridflow.exports import SurveyCSVExporter
from gridflow.exports.excel_exporter import POLES_COLUMNS
from gridflow.exports.pdf_exporter import SurveyPDFExporter

PYTHON3 = "python3"


def _run_python3(*args: str) -> subprocess.CompletedProcess[str]:
    return subprocess.run(
        [PYTHON3, *args],
        capture_output=True,
        text=True,
        check=True,
    )


def _write_notes(
    path: Path, support: str, fid: str, spn: str, pole_type: str = "Intermediate"
) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        "\n".join(
            [
                f"# Pole {support}",
                "",
                "## Identity",
                f"Support number: {support}",
                f"ENWL FID: {fid}",
                f"SPN: {spn}",
                f"Pole type: {pole_type}",
                "Pole class: Single Wood Pole",
                "Support diameter: Stout",
                "Coordinates from ENWL: 54.1000, -2.7000",
                "",
                "## Conductor record 1",
                "- fid: 7001",
                "- feature_type: HV Conductor",
                "- voltage: 11kV",
                "- material: Aluminium",
                "- cable_size: 50mm2",
                "- text_map: 3x 50 Al 11",
            ]
        ),
        encoding="utf-8",
    )


def _make_survey(tmp_path: Path, with_photos: bool = True) -> tuple[Path, Path]:
    survey_root = tmp_path / "P_TEST"
    poles_root = survey_root / "enwl_enrichment_clean"
    pole = poles_root / "01_SUPPORT_123456"
    (pole / "field_photos").mkdir(parents=True, exist_ok=True)
    if with_photos:
        (pole / "field_photos" / "IMG_0001.jpg").write_bytes(b"img")
    _write_notes(pole / "notes" / "pole_notes.md", "123456", "1001", "SPN001")

    trace_path = survey_root / "trace.geojson"
    trace_path.write_text(
        json.dumps(
            {
                "type": "FeatureCollection",
                "features": [
                    {
                        "type": "Feature",
                        "properties": {
                            "FID": "1001",
                            "feature_type": "Pole",
                            "support_no": "123456",
                            "spn": "SPN001",
                            "pole_type": "Intermediate",
                            "pole_class": "Single Wood Pole",
                            "support_diameter": "Stout",
                        },
                        "geometry": {"type": "Point", "coordinates": [-2.7, 54.1]},
                    }
                ],
            }
        ),
        encoding="utf-8",
    )
    return survey_root, trace_path


def test_excel_export_creates_four_sheets(tmp_path):
    survey_root, trace_path = _make_survey(tmp_path)
    output = tmp_path / "survey.xlsx"
    _run_python3(
        "scripts/export_survey_excel.py",
        "--survey",
        str(survey_root),
        "--trace",
        str(trace_path),
        "--output",
        str(output),
    )

    probe = _run_python3(
        "-c",
        (
            "from openpyxl import load_workbook; "
            f"wb=load_workbook(r'{output}'); "
            "print('|'.join(wb.sheetnames))"
        ),
    )
    assert probe.stdout.strip().split("|") == ["Summary", "Poles", "Conflicts", "Photos"]


def test_poles_sheet_has_correct_columns(tmp_path):
    survey_root, trace_path = _make_survey(tmp_path)
    output = tmp_path / "survey.xlsx"
    _run_python3(
        "scripts/export_survey_excel.py",
        "--survey",
        str(survey_root),
        "--trace",
        str(trace_path),
        "--output",
        str(output),
    )

    probe = _run_python3(
        "-c",
        (
            "from openpyxl import load_workbook; "
            f"wb=load_workbook(r'{output}'); "
            "print('|'.join(str(cell.value) for cell in wb['Poles'][1]))"
        ),
    )
    assert probe.stdout.strip().split("|") == POLES_COLUMNS


def test_csv_export_creates_header_row(tmp_path):
    survey_root, trace_path = _make_survey(tmp_path)
    output = tmp_path / "survey.csv"
    SurveyCSVExporter().export(survey_root, trace_path, output)

    with output.open(encoding="utf-8") as handle:
        rows = list(csv.reader(handle))
    assert rows[0] == POLES_COLUMNS
    assert len(rows) == 2


def test_csv_has_correct_column_count(tmp_path):
    survey_root, trace_path = _make_survey(tmp_path)
    output = tmp_path / "survey.csv"
    SurveyCSVExporter().export(survey_root, trace_path, output)

    with output.open(encoding="utf-8") as handle:
        header = next(csv.reader(handle))
    assert len(header) == 18


def test_missing_trace_path_handled_gracefully(tmp_path):
    survey_root, _trace_path = _make_survey(tmp_path)
    output = tmp_path / "missing-trace.xlsx"
    _run_python3(
        "scripts/export_survey_excel.py",
        "--survey",
        str(survey_root),
        "--trace",
        str(tmp_path / "no-trace.geojson"),
        "--output",
        str(output),
    )
    assert output.exists()
    csv_output = tmp_path / "missing-trace.csv"
    bundle = SurveyCSVExporter().export(survey_root, tmp_path / "no-trace.geojson", csv_output)
    assert bundle.total_poles == 1
    assert bundle.pole_rows[0]["conflict_count"] == 0


def test_missing_photos_handled_gracefully(tmp_path):
    survey_root, trace_path = _make_survey(tmp_path, with_photos=False)
    output = tmp_path / "survey.csv"
    bundle = SurveyCSVExporter().export(survey_root, trace_path, output)
    assert bundle.pole_rows[0]["photo_count"] == 0


def test_missing_conflicts_handled_gracefully(tmp_path):
    survey_root, trace_path = _make_survey(tmp_path)
    output = tmp_path / "survey.csv"
    bundle = SurveyCSVExporter().export(survey_root, trace_path, output)
    assert bundle.pole_rows[0]["conflict_count"] == 0


def test_empty_survey_folder_raises_clear_error(tmp_path):
    survey_root = tmp_path / "EMPTY"
    (survey_root / "enwl_enrichment_clean").mkdir(parents=True)

    try:
        SurveyCSVExporter().export(survey_root, tmp_path / "trace.geojson", tmp_path / "out.csv")
    except ValueError as exc:
        assert "no pole folders" in str(exc).lower()
    else:
        raise AssertionError("Expected ValueError for empty survey")


def test_pdf_export_creates_file_with_multiple_pages(tmp_path):
    survey_root = tmp_path / "P_TEST"
    poles_root = survey_root / "enwl_enrichment_clean"
    for idx in range(1, 14):
        pole = poles_root / f"{idx:02d}_SUPPORT_{100000 + idx}"
        (pole / "field_photos").mkdir(parents=True, exist_ok=True)
        (pole / "field_photos" / "IMG_0001.jpg").write_bytes(b"img")
        _write_notes(
            pole / "notes" / "pole_notes.md",
            str(100000 + idx),
            str(1000 + idx),
            f"SPN{idx:03d}",
        )

    trace_path = survey_root / "trace.geojson"
    trace_path.write_text(
        json.dumps(
            {
                "type": "FeatureCollection",
                "features": [
                    {
                        "type": "Feature",
                        "properties": {
                            "FID": str(1000 + idx),
                            "feature_type": "Pole",
                            "support_no": str(100000 + idx),
                            "spn": f"SPN{idx:03d}",
                            "pole_type": "Intermediate",
                            "pole_class": "Single Wood Pole",
                            "support_diameter": "Stout",
                        },
                        "geometry": {"type": "Point", "coordinates": [-2.7, 54.1]},
                    }
                    for idx in range(1, 14)
                ],
            }
        ),
        encoding="utf-8",
    )

    output = tmp_path / "survey.pdf"
    bundle = SurveyPDFExporter().export(survey_root, trace_path, output)
    assert bundle.total_poles == 13
    assert output.exists()
    pdf_bytes = output.read_bytes()
    assert len(pdf_bytes) > 0
    assert pdf_bytes.count(b"/Type /Page") >= 2
